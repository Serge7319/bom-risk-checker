import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from src.health_score import calculate_bom_health_score, generate_executive_summary


def save_results_to_csv(results: list, output_path: str) -> None:
    df = pd.DataFrame(results)
    df.to_csv(output_path, index=False)


def save_results_to_excel(results: list, output_path: str) -> None:
    df = pd.DataFrame(results)

    high_risk_df = df[df["Risk Level"] == "High"]

    health_data = calculate_bom_health_score(df)
    executive_bullets = generate_executive_summary(df)

    summary_rows = [
        ["BOM Health Score", f"{health_data['health_score']} / 100"],
        ["Overall Status", health_data["health_status"]],
        ["Summary Message", health_data["summary_message"]],
        ["", ""],
        ["Metric", "Value"],
        ["Total Parts", len(df)],
        ["High Risk Parts", len(df[df["Risk Level"] == "High"])],
        ["Medium Risk Parts", len(df[df["Risk Level"] == "Medium"])],
        ["Low Risk Parts", len(df[df["Risk Level"] == "Low"])],
        ["Obsolete / EOL Parts", len(df[df["Lifecycle Status"].isin(["Obsolete", "EOL"])] )],
        ["Unknown Lifecycle Parts", len(df[df["Lifecycle Status"] == "Unknown"])],
        ["", ""],
        ["Executive Summary", ""],
    ]

    for bullet in executive_bullets:
        summary_rows.append([bullet, ""])

    summary_df = pd.DataFrame(summary_rows, columns=["Category", "Result"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        df.to_excel(writer, sheet_name="Detailed Results", index=False)
        high_risk_df.to_excel(writer, sheet_name="High Risk Parts", index=False)

    format_excel_report(output_path)


def format_excel_report(output_path: str) -> None:
    workbook = load_workbook(output_path)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="FECACA")
    medium_fill = PatternFill("solid", fgColor="FEF3C7")
    low_fill = PatternFill("solid", fgColor="DCFCE7")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        autofit_columns(sheet)

    if "Summary" in workbook.sheetnames:
        format_summary_sheet(workbook["Summary"])

    if "Detailed Results" in workbook.sheetnames:
        apply_risk_formatting(workbook["Detailed Results"], high_fill, medium_fill, low_fill)

    if "High Risk Parts" in workbook.sheetnames:
        apply_risk_formatting(workbook["High Risk Parts"], high_fill, medium_fill, low_fill)

    workbook.save(output_path)


def apply_risk_formatting(sheet, high_fill, medium_fill, low_fill):
    headers = [cell.value for cell in sheet[1]]

    if "Risk Level" not in headers:
        return

    risk_col_index = headers.index("Risk Level") + 1

    for row in sheet.iter_rows(min_row=2):
        risk_value = row[risk_col_index - 1].value

        if risk_value == "High":
            fill = high_fill
        elif risk_value == "Medium":
            fill = medium_fill
        elif risk_value == "Low":
            fill = low_fill
        else:
            continue

        for cell in row:
            cell.fill = fill


def autofit_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        adjusted_width = min(max_length + 2, 45)
        sheet.column_dimensions[column_letter].width = adjusted_width


def format_summary_sheet(sheet):
    """
    Makes the Summary tab look like an executive dashboard.
    """

    # Sheet title
    sheet.insert_rows(1, 2)
    sheet["A1"] = "BOM Risk Checker Report"
    sheet["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="111827")
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A1:B1")

    sheet["A2"] = "Executive Summary"
    sheet["A2"].font = Font(size=12, bold=True, color="374151")
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:B2")

    # Format key dashboard rows
    for row in range(4, sheet.max_row + 1):
        category = sheet[f"A{row}"].value

        if category == "BOM Health Score":
            sheet[f"A{row}"].font = Font(bold=True, size=14)
            sheet[f"B{row}"].font = Font(bold=True, size=18, color="2563EB")

        elif category == "Overall Status":
            sheet[f"A{row}"].font = Font(bold=True)
            sheet[f"B{row}"].font = Font(bold=True, color="FFFFFF")

            status = str(sheet[f"B{row}"].value)

            if status == "Healthy":
                sheet[f"B{row}"].fill = PatternFill("solid", fgColor="16A34A")
            elif status == "Moderate Risk":
                sheet[f"B{row}"].fill = PatternFill("solid", fgColor="F59E0B")
            elif status == "High Risk":
                sheet[f"B{row}"].fill = PatternFill("solid", fgColor="DC2626")

        elif category == "Metric":
            sheet[f"A{row}"].font = Font(bold=True, color="FFFFFF")
            sheet[f"B{row}"].font = Font(bold=True, color="FFFFFF")
            sheet[f"A{row}"].fill = PatternFill("solid", fgColor="374151")
            sheet[f"B{row}"].fill = PatternFill("solid", fgColor="374151")

        elif category == "Executive Summary":
            sheet[f"A{row}"].font = Font(bold=True, size=14, color="111827")
            sheet[f"A{row}"].fill = PatternFill("solid", fgColor="E5E7EB")
            sheet[f"B{row}"].fill = PatternFill("solid", fgColor="E5E7EB")

    # Better sizing
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 35

    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 24